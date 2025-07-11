import json
import openpyxl
import os
from datetime import datetime
from da_api_client import DAAPIClient
from selenium_bot import DASeleniumBot
import config_reader

class PatientCreator:
    def __init__(self):
        self.da_api = DAAPIClient()
        self.selenium_bot = DASeleniumBot()
        self.config = config_reader.read_config()
    
    def process_failed_patients(self, patient_list, helper_id, start_date=None, end_date=None, dry_run=False, api_available=True):
        """
        Main workflow for processing patients that failed order upload
        
        Args:
            patient_list: List of patient names that need to be created
            helper_id: Helper ID for impersonation
            start_date: Start date for search (optional)
            end_date: End date for search (optional)
            dry_run: If True, test workflow without creating patients
            api_available: If True, DA API credentials are available
        """
        try:
            print(f"Starting patient creation workflow for {len(patient_list)} patients")
            if dry_run:
                print("DRY RUN MODE: Testing workflow without patient creation")
            if not api_available:
                print("API MODE: Limited functionality - document search only")
            
            # Step 1: Setup Selenium bot and login
            self.selenium_bot.setup_driver()
            self.selenium_bot.login_to_da()
            self.selenium_bot.navigate_to_search(helper_id)
            
            # Step 2: Fetch 485 certificates for target patients
            certificates_485 = self.selenium_bot.fetch_485_certificates(
                patient_names=patient_list,
                start_date=start_date,
                end_date=end_date
            )
            
            if not certificates_485:
                print("No 485 certificates found for the specified patients")
                return []
            
            print(f"Found {len(certificates_485)} 485 certificates to process")
            
            # Step 3: Process each certificate and create patients
            created_patients = []
            
            for cert in certificates_485:
                try:
                    print(f"Processing 485 certificate for patient: {cert['patient_name']}")
                    
                    # Initialize result object
                    result = {
                        'patient_name': cert['patient_name'],
                        'doc_id': cert['doc_id'],
                        'creation_status': 'Pending',
                        'creation_message': 'Processing...',
                        'patient_payload': None
                    }
                    
                    if api_available:
                        # Get detailed document information using DA API
                        doc_details = self.da_api.get_document_by_id(cert['doc_id'])
                        
                        # Extract patient information from document
                        patient_info = self._extract_patient_info_from_doc(doc_details, cert)
                        
                        # Download the 485 PDF for further processing if needed
                        download_folder = os.path.join(self.config['REPORTS_PATH'], "485_certificates")
                        file_path, doc_metadata = self.da_api.get_document_file(cert['doc_id'], download_folder)
                        patient_info['pdf_path'] = file_path
                        
                        # Create comprehensive patient payload
                        patient_payload = self._create_patient_payload(patient_info, doc_metadata)
                        result['patient_payload'] = patient_payload
                        
                        if dry_run:
                            result['creation_status'] = 'Dry Run Success'
                            result['creation_message'] = 'Patient payload created successfully (dry run)'
                        else:
                            # Attempt to create patient
                            creation_result = self._create_patient_in_system(patient_payload)
                            result['creation_status'] = creation_result['status']
                            result['creation_message'] = creation_result['message']
                    else:
                        # Limited mode - just document search
                        result['creation_status'] = 'Limited Mode'
                        result['creation_message'] = 'Document found, but API not available for patient creation'
                    
                    created_patients.append(result)
                    
                except Exception as e:
                    print(f"Error processing patient {cert['patient_name']}: {str(e)}")
                    created_patients.append({
                        'patient_name': cert['patient_name'],
                        'doc_id': cert.get('doc_id', 'Unknown'),
                        'creation_status': 'Failed',
                        'creation_message': str(e),
                        'patient_payload': None
                    })
            
            # Step 4: Generate report
            self._generate_patient_creation_report(created_patients)
            
            return created_patients
            
        except Exception as e:
            print(f"Error in patient creation workflow: {str(e)}")
            raise
        finally:
            self.selenium_bot.close()
    
    def _extract_patient_info_from_doc(self, doc_details, cert_info):
        """Extract patient information from document details"""
        patient_info = {
            'name': cert_info['patient_name'],
            'doc_id': cert_info['doc_id'],
            'doc_type': cert_info['doc_type'],
            'order_date': cert_info['order_date'],
            'physician': cert_info['physician']
        }
        
        try:
            # Extract additional info from DA document response
            if 'value' in doc_details and doc_details['value']:
                doc_value = doc_details['value']
                
                # Get patient ID from document to fetch more details
                if 'patientId' in doc_value:
                    patient_id = doc_value['patientId']
                    patient_info['da_patient_id'] = patient_id
                    
                    # Fetch detailed patient information
                    try:
                        patient_details = self.da_api.get_patient_by_id(patient_id)
                        if 'value' in patient_details:
                            patient_data = patient_details['value']
                            patient_info.update(self._extract_patient_details(patient_data))
                    except Exception as e:
                        print(f"Could not fetch patient details for ID {patient_id}: {str(e)}")
                
                # Extract other relevant document information
                if 'physicianSigndate' in doc_value:
                    patient_info['physician_sign_date'] = doc_value['physicianSigndate']
                
                if 'status' in doc_value:
                    status = doc_value['status']
                    patient_info.update(self._extract_status_info(status))
        
        except Exception as e:
            print(f"Error extracting patient info from document: {str(e)}")
        
        return patient_info
    
    def _extract_patient_details(self, patient_data):
        """Extract patient details from DA patient API response"""
        details = {}
        
        if 'patientInfo' in patient_data:
            info = patient_data['patientInfo']
            details.update({
                'first_name': info.get('firstName', ''),
                'middle_name': info.get('middleInitial', ''),
                'last_name': info.get('lastName', ''),
                'full_name': info.get('name', ''),
                'dob': info.get('dob', ''),
                'sex': info.get('sex', ''),
                'address': info.get('address', ''),
                'city': info.get('city', ''),
                'state': info.get('state', ''),
                'zip_code': info.get('zipCode', ''),
                'phone': info.get('phone', ''),
                'ssn': info.get('ssn', ''),
                'mrn': info.get('medicalRecordNumber', ''),
                'insurance_number': info.get('insuranceNumber', ''),
                'pay_source': info.get('paySource', ''),
                'physician_npi': info.get('physicianNpi', ''),
                'physician_name': info.get('physicianName', '')
            })
        
        if 'patientStatus' in patient_data:
            status = patient_data['patientStatus']
            details.update(self._extract_status_info(status))
        
        return details
    
    def _extract_status_info(self, status_data):
        """Extract status information from patient status"""
        status_info = {}
        
        if isinstance(status_data, dict):
            status_info.update({
                'start_of_care_date': status_data.get('startOfCareDate', ''),
                'cert_period_from': status_data.get('certPeriodFrom', ''),
                'cert_period_to': status_data.get('certPeriodTo', ''),
                'prognosis': status_data.get('prognosis', ''),
                'state': status_data.get('state', ''),
                'medications': status_data.get('medications', ''),
                'allergies': status_data.get('allergies', ''),
                'goals': status_data.get('goals', ''),
                'orders': status_data.get('orders', '')
            })
            
            # Extract diagnoses if available
            if 'diagnoses' in status_data and status_data['diagnoses']:
                diagnoses = []
                for diag in status_data['diagnoses']:
                    diagnoses.append({
                        'description': diag.get('description', ''),
                        'code': diag.get('code', ''),
                        'type': diag.get('diagnosisType', ''),
                        'date': diag.get('date', '')
                    })
                status_info['diagnoses'] = diagnoses
        
        return status_info
    
    def _create_patient_payload(self, patient_info, doc_metadata):
        """Create comprehensive patient payload for creation"""
        # Parse patient name
        full_name = patient_info.get('name', patient_info.get('full_name', ''))
        first_name = patient_info.get('first_name', '')
        middle_name = patient_info.get('middle_name', '')
        last_name = patient_info.get('last_name', '')
        
        # If names are not extracted, try to parse from full name
        if not first_name and full_name:
            name_parts = full_name.split(', ')
            if len(name_parts) >= 2:
                last_name = name_parts[0].strip()
                remaining = name_parts[1].strip().split(' ')
                first_name = remaining[0] if remaining else ''
                middle_name = remaining[1] if len(remaining) > 1 else ''
        
        # Get dates in proper format
        start_of_care = self._format_date(patient_info.get('start_of_care_date'))
        cert_from = self._format_date(patient_info.get('cert_period_from'))
        cert_to = self._format_date(patient_info.get('cert_period_to'))
        dob = self._format_date(patient_info.get('dob'))
        
        # Create the patient payload structure
        payload = {
            "patientInfo": {
                "name": full_name,
                "fullAddress": None,
                "id": 0,
                "firstName": first_name,
                "middleInitial": middle_name,
                "lastName": last_name,
                "sex": patient_info.get('sex', ''),
                "dob": dob,
                "suite": None,
                "address": patient_info.get('address'),
                "phone": patient_info.get('phone'),
                "state": patient_info.get('state'),
                "city": patient_info.get('city'),
                "zipCode": patient_info.get('zip_code'),
                "paySource": patient_info.get('pay_source'),
                "insuranceNumber": patient_info.get('insurance_number', '0000'),
                "medicalRecordNumber": patient_info.get('mrn'),
                "ssn": patient_info.get('ssn'),
                "physicianHelperId": 0,
                "externalId": None,
                "status": None,
                "physicianName": patient_info.get('physician_name', ''),
                "physicianNpi": patient_info.get('physician_npi', '')
            },
            "patientStatus": {
                "medications": patient_info.get('medications'),
                "dmeAndSupplies": None,
                "safetyMeasures": None,
                "nutritionalRequirements": None,
                "allergies": patient_info.get('allergies'),
                "functionalLimitations": {
                    "amputation": True,
                    "bowel": True,
                    "contracture": True,
                    "hearing": True,
                    "paralysis": True,
                    "endurance": True,
                    "ambulation": True,
                    "speech": True,
                    "legallyBlind": True,
                    "dyspnea": True,
                    "others": None
                },
                "activitiesPermitted": {
                    "completeBedrest": True,
                    "bedrestBRP": True,
                    "upAsTolerated": True,
                    "transferBedChair": True,
                    "exercisesPrescribed": True,
                    "partialWeightBearing": True,
                    "independentAtHome": True,
                    "crutches": True,
                    "cane": True,
                    "wheelChair": True,
                    "walker": True,
                    "noRestrictions": True,
                    "others": None
                },
                "mentalStatus": {
                    "oriented": True,
                    "forgetful": True,
                    "comatose": True,
                    "depressed": True,
                    "disoriented": True,
                    "lethargic": True,
                    "agitated": True,
                    "others": None
                },
                "prognosis": patient_info.get('prognosis', 'poor'),
                "orders": patient_info.get('orders'),
                "goals": patient_info.get('goals'),
                "state": "admitted",
                "startOfCareDate": start_of_care,
                "certPeriodFrom": cert_from,
                "certPeriodTo": cert_to,
                "diagnoses": patient_info.get('diagnoses', [
                    {
                        "description": "string",
                        "code": "string",
                        "eorO": "1",
                        "diagnosisType": "principal",
                        "date": datetime.now().strftime("%Y-%m-%dT%H:%M:%S.%fZ")
                    }
                ])
            },
            "physicianNpi": patient_info.get('physician_npi', ''),
            "clinicianId": {
                "npi": "NULL",
                "id": self.da_api.api_config['clinician_id'],
                "externalId": "NULL"
            },
            "careProviderId": {
                "npi": "NULL",
                "id": self.da_api.api_config['caretaker_id'],
                "externalId": "NULL"
            }
        }
        
        return payload
    
    def _format_date(self, date_str):
        """Format date string to expected format"""
        if not date_str:
            return None
        
        try:
            # Try different date formats
            for fmt in ['%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            
            # If no format matches, return as is
            return date_str
        except Exception:
            return date_str
    
    def _create_patient_in_system(self, patient_payload):
        """Create patient in the system using DA API"""
        try:
            response = self.da_api.create_patient(patient_payload)
            
            if response.get('isSuccess'):
                return {
                    'status': 'Success',
                    'message': f"Patient created successfully: {response['value']['actionType']}:{response['value']['id']}"
                }
            else:
                return {
                    'status': 'Failed',
                    'message': response.get('errorMessage', 'Unknown error')
                }
        
        except Exception as e:
            return {
                'status': 'Failed',
                'message': str(e)
            }
    
    def _generate_patient_creation_report(self, created_patients):
        """Generate Excel report of patient creation results"""
        try:
            os.makedirs(self.config['REPORTS_PATH'], exist_ok=True)
            
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Patient Creation Report"
            
            # Headers
            headers = [
                'Patient Name', 'Document ID', 'Creation Status', 'Creation Message',
                'First Name', 'Last Name', 'DOB', 'MRN', 'Physician NPI',
                'Start of Care', 'Cert Period From', 'Cert Period To'
            ]
            
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Data rows
            for row, patient in enumerate(created_patients, 2):
                payload = patient.get('patient_payload', {})
                patient_info = payload.get('patientInfo', {}) if payload else {}
                patient_status = payload.get('patientStatus', {}) if payload else {}
                
                worksheet.cell(row=row, column=1, value=patient['patient_name'])
                worksheet.cell(row=row, column=2, value=patient['doc_id'])
                worksheet.cell(row=row, column=3, value=patient['creation_status'])
                worksheet.cell(row=row, column=4, value=patient['creation_message'])
                worksheet.cell(row=row, column=5, value=patient_info.get('firstName', ''))
                worksheet.cell(row=row, column=6, value=patient_info.get('lastName', ''))
                worksheet.cell(row=row, column=7, value=patient_info.get('dob', ''))
                worksheet.cell(row=row, column=8, value=patient_info.get('medicalRecordNumber', ''))
                worksheet.cell(row=row, column=9, value=patient_info.get('physicianNpi', ''))
                worksheet.cell(row=row, column=10, value=patient_status.get('startOfCareDate', ''))
                worksheet.cell(row=row, column=11, value=patient_status.get('certPeriodFrom', ''))
                worksheet.cell(row=row, column=12, value=patient_status.get('certPeriodTo', ''))
            
            # Save report
            report_filename = f"patient_creation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_path = os.path.join(self.config['REPORTS_PATH'], report_filename)
            workbook.save(report_path)
            
            print(f"Patient creation report saved: {report_path}")
            
        except Exception as e:
            print(f"Error generating report: {str(e)}") 