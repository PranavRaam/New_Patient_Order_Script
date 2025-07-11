from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
from openpyxl import load_workbook
import time
from datetime import datetime, timedelta
import os
import pandas as pd
import json
from datetime import datetime
import shutil
import ReadConfig as rc
import SplitPDF
import CommonUtil as cu
import openpyxl
def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def sendSigned(url,email,password,credName,reportFolderName):
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/OrderTemplate.xlsx"
    order_template = order_template.replace('\\', '/')

    driver = webdriver.Chrome()
    driver.maximize_window()
    try:
        for ctr in range(3):
            try:
                cu.login_to_EHR(url,email,password,credName,driver)
                view_span = wait_and_find_element(driver, By.XPATH, "//span[contains(text(), 'View')]")
                actions = ActionChains(driver)
                actions.move_to_element(view_span).perform()
                print("View Span Clicked")

                orders_management_span = wait_and_find_element(driver, By.XPATH, "//span[contains(text(), 'Orders Management')]")
                actions = ActionChains(driver)
                actions.move_to_element(orders_management_span).perform()
                print("Orders Management Span Clicked")
                
                orders_to_be_sent_div = wait_and_find_element(driver, By.XPATH, "//div[contains(text(), 'Orders To Be Sent')]")
                click_element(orders_to_be_sent_div)
                time.sleep(10)
                

                destination_package = openpyxl.load_workbook(order_template)
                worksheet = destination_package.active
                end_row = worksheet.max_row
                for row in range(2, end_row + 1):
                    status= cu.clean_null_data(worksheet.cell(row, 20).value)
                    #Status Changed
                    if cu.clean_null_data(worksheet.cell(row, 14).value) and "Status Changed" not in status:
                        order_number = worksheet.cell(row, 2).value
                        try:
                            search_input = driver.find_element(By.CLASS_NAME, "grid-search")
                            search_input.clear()
                            search_input.send_keys(order_number)
                            time.sleep(5)

                            tr = driver.find_element(By.CLASS_NAME, 'match')
                            if tr:
                                checkbox = driver.find_element(By.ID, 'OrdersManagement_Order_CheckAll')
                                if checkbox.is_selected():
                                    checkbox.click()
                                checkbox.click()
                                print("Checkbox checked.")
                                time.sleep(2)  

                                #SEND
                                send_button = driver.find_element(By.ID, "OrdersManagement_List_OrdersToBeSent_SendButton")
                                driver.execute_script("arguments[0].click();", send_button)
                                worksheet.cell(row, 20).value = "Status Changed"

                            else:
                                time.sleep(2)  
                
                        except NoSuchElementException:
                            print("No orders found. Continue with the next steps.")
                            continue

                destination_package.save(order_template)
                driver.quit()
            except Exception as ex:
                print(str(ex))
    except Exception as ex:
                print(str(ex))
        
# sendSigned("https://accounts.axxessweb.com/Login","cliffard@doctoralliance.com","Dallas@1234","Standards Home Care","Axxess-StandardHomeHealth")