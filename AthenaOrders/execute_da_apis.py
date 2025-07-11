import requests
import ReadConfig as rc
from datetime import datetime
import CommonUtil as cu
import openpyxl
import json
import os
import base64
import re

def execute_da_signed_order_download(access_token, daAPIPatientUrl, reportFolderName, daAPITokenCaretakerId):
    base_folder = cu.getFolderPath("O", reportFolderName)
    working_file=os.path.join(base_folder,"OrderTemplate.xlsx")
    signed_orders_folder = os.path.join(base_folder, "SignedOrders")
    working_file = working_file.replace('\\', '/')
    signed_orders_folder = signed_orders_folder.replace('\\', '/')
    try:
        destination_package = openpyxl.load_workbook(working_file)
        destination_worksheet = destination_package.active
        signed_date=""
        start_of_care=""
        cert_period_from=""
        cert_period_to=""
        for row in range(2, destination_worksheet.max_row + 1):
            try:
                doc_id=cu.clean_null_data(destination_worksheet.cell(row, 14).value)
                signed_by_physician = cu.clean_null_data(destination_worksheet.cell(row, 12).value)
                signed_by_physician_date = cu.clean_null_data(destination_worksheet.cell(row, 9).value)
                if doc_id and signed_by_physician and not signed_by_physician_date:
                    signed_date,start_of_care,cert_period_from,cert_period_to = get_Signed_pdf(access_token, daAPIPatientUrl, doc_id, signed_orders_folder, daAPITokenCaretakerId)
                    destination_worksheet.cell(row, 9).value=signed_date
                    destination_worksheet.cell(row, 24).value=start_of_care
                    destination_worksheet.cell(row, 5).value=cert_period_from
                    destination_worksheet.cell(row, 6).value=cert_period_to
            except Exception as e:
                destination_worksheet.cell(row, 20).value=str(e)
                continue
                
        destination_package.save(working_file)
    except Exception as e:
        raise Exception("Error in Downloading signed orders: "+str(e))



def get_Signed_pdf(access_token, daAPIPatientUrl, doc_id, signed_orders_folder, daAPITokenCaretakerId):
    signed_date=""
    start_of_care=""
    cert_period_from=""
    cert_period_to=""

    url= daAPIPatientUrl.split('patient')[0]+"document/getfile?docId.id="+doc_id
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + access_token
    }
    data = {
    "onlyUnfiled": True,
    "careProviderId": {
        "id": daAPITokenCaretakerId,
        "externalId": ""
    },
    "dateFrom": None,
    "dateTo": None,
    "page": 1,
    "recordsPerPage": 10
    }
    response = requests.get(url, headers=headers, json=data)

    if response.status_code == 200:
        response_json = json.loads(response.content)
        # data = response.json()
        document_buffer = response_json['value']['documentBuffer']
        try:
            start_of_care = cu.date_in_standard_format(response_json['value']['document']['status']['startOfCareDate'])
            cert_period_from = cu.date_in_standard_format(response_json['value']['document']['status']['certPeriodFrom'])
            cert_period_to = cu.date_in_standard_format(response_json['value']['document']['status']['certPeriodTo'])
        except Exception as e:
            pass
        document_data_bytes = base64.b64decode(document_buffer)
        destination_path=os.path.join(signed_orders_folder,f"{doc_id}.pdf")
        destination_path = destination_path.replace('\\', '/')
        with open(destination_path, "wb") as f:
            f.write(document_data_bytes)
        signed_date=get_signed_date(access_token, daAPIPatientUrl, doc_id)
    return signed_date,start_of_care,cert_period_from,cert_period_to


def get_signed_date(access_token, daAPIPatientUrl, doc_id):
    signed_date=""
    url= daAPIPatientUrl.split('patient')[0]+"document/get?docId.id="+doc_id
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + access_token
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        response_json = json.loads(response.content)
        try:
            signed_date=response_json["value"]["physicianSigndate"]
            if signed_date:
                signed_date=cu.date_in_standard_format(signed_date)
        except Exception as e:
            pass
    return signed_date
    





    
# execute_da_patient_creation("Axxess-StandardHomeHealth")