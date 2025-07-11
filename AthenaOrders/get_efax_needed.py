import requests
import json
import ReadConfig as rc
import CommonUtil as cu
import os
import openpyxl

def get_efax_needed(reportFolderName):
    configuration={}
    configuration=rc.readConfig()
    api_key=configuration["APIKey"]
    api_url=configuration["APIBaseURL"]
    headers = {}
    headers["X-SERVICE-KEY"] = api_key
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/OrderTemplate.xlsx"
    order_template = order_template.replace('\\', '/')
    download_folder=configuration["DownloadPath"]
    template_folder=configuration["OrderTemplatePath"]
    # shutil.copy2(template_folder, order_template)
    # try:
    #     destination_package = openpyxl.load_workbook(working_file)
    #     destination_worksheet = destination_package.active
    #     for row in range(2, destination_worksheet.max_row + 1):
    #         try:
    #             doc_id=cu.clean_null_data(destination_worksheet.cell(row, 14).value)
    #             signed_by_physician = cu.clean_null_data(destination_worksheet.cell(row, 12).value)
    #             order_no = cu.clean_null_data(destination_worksheet.cell(row, 1).value)
    #             if signed_by_physician and not order_no:
    #                 endpoint = f"/api/Order/GetOrdersByDocNo/{doc_id}"
    #                 api_endpoint = api_url + endpoint
    #                 order_data = {}
    #                 response = requests.get(api_endpoint, headers=headers)

    #                 if response.status_code == 200:
    #                     data = response.json()
    #                     json_str = json.dumps(data)
    #                     if json_str:
    #                         order_data = json.loads(json_str)
    #                         # print(order_data)
    #                         destination_worksheet.cell(row, 1).value= order_data["id"]
    #                         destination_worksheet.cell(row, 2).value= order_data["orderNo"]
    #                         destination_worksheet.cell(row, 4).value= cu.get_date_string(order_data["orderDate"])
    #                         destination_worksheet.cell(row, 5).value= cu.get_date_string(order_data["episodeStartDate"])
    #                         destination_worksheet.cell(row, 6).value= cu.get_date_string(order_data["episodeEndDate"])
    #                         destination_worksheet.cell(row, 7).value= order_data["mrn"]
    #                         destination_worksheet.cell(row, 8).value= cu.get_date_string(order_data["sentToPhysicianDate"])

    #         except Exception as e:
    #             destination_worksheet.cell(row, 20).value=str(e)

    #     destination_package.save(working_file)
    # except Exception as e:
    #     print(e)

    
# get_order_info("Axxess-StandardHomeHealth")

