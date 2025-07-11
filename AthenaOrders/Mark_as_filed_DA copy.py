import openpyxl
import json
import time
import ReadConfig as rc
import os
import re
import FetchAthenaConfig as fc
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import calendar
import pyautogui
import CommonUtil as cu
import shutil
import pandas as pd

def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def click_patient_select_element(driver):
    patient_select_element = wait_and_find_element(driver, By.CSS_SELECTOR, "span.select2-selection__rendered span.select2-selection__placeholder")
    patient_select_element.click()
    time.sleep(2)

def interact_with_search_field(driver, search_term):
    input_css_selector = "input.select2-search__field"
    search_input = wait_and_find_element(driver, By.CSS_SELECTOR, input_css_selector)
    search_input.clear()
    search_input.send_keys(search_term)
    time.sleep(10)

def is_document_to_be_filed(document_id,worksheet):
    excel_row = worksheet.max_row+1
    flag=False
    for r in range(2, excel_row):
        try:
            doc_id=cu.clean_null_data(worksheet.cell(r, 14).value)
            signed_order_upload_status=cu.clean_null_data(worksheet.cell(r, 13).value)
            if doc_id==document_id and signed_order_upload_status:
                flag=True
                break
        except Exception as e:
            continue
    return flag



def mark_as_filed(da_url, da_login, da_password, reportFolderName, helper_id):
    configuration = rc.readConfig()
    rpaName = configuration["RPA"]
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/OrderTemplate.xlsx"
    order_template = order_template.replace('\\', '/')
    signed_orders_folder = os.path.join(working_folder, "SignedOrders")
    signed_orders_folder = signed_orders_folder.replace('\\', '/')
    df = pd.read_excel(order_template)
    filtered_df = df[df['Uploaded Signed Order Status'] == 'TRUE']
    
    driver = webdriver.Chrome()
    driver.maximize_window()
    try:
        cu.login_to_da(da_url, da_login, da_password, driver)
        driver.get("https://backoffice.doctoralliance.com/Search")
        time.sleep(7)
        query_input = driver.find_element(By.ID, "Query")
        query_input.send_keys(helper_id)
        time.sleep(2)
        search_type_dropdown = driver.find_element(By.ID, "select2-SearchType-container")
        search_type_dropdown.click()
        time.sleep(1)
        input_field = driver.find_element(By.CLASS_NAME, "select2-search__field")
        input_field.send_keys("Users")
        time.sleep(3)
        wait = WebDriverWait(driver,10)
        first_result = wait.until(EC.visibility_of_element_located((By.XPATH, "//li[contains(@id, 'select2-SearchType-result')][1]")))
        first_result.click()
        time.sleep(3)
        button = driver.find_element(By.CLASS_NAME, "btn-success")
        button.click()
        time.sleep(5)
        row = driver.find_element(By.CLASS_NAME, "linkedRow")
        row.click()
        time.sleep(5)

        link = driver.find_element(By.LINK_TEXT,"Impersonate")
        link.click()
        time.sleep(7)

        driver.switch_to.window(driver.window_handles[1])

        signed_link = wait_and_find_element(driver, By.XPATH, "//a[contains(@href, '/Documents/Signed')]")
        signed_link.click()
        time.sleep(2) 

        start_date_input = wait_and_find_element(driver, By.ID, "StartDatePicker")
        start_date_input.clear()
        start_date=(datetime.now() - timedelta(days=4)).strftime("%m/%d/%Y")
        start_date_input.send_keys(cu.get_date_string(start_date))
        time.sleep(1) 

        # Find the end date input field and fill it with the current date
        end_date_input = wait_and_find_element(driver, By.ID, "EndDatePicker")
        end_date_input.clear()
        end_date=datetime.now().strftime("%m/%d/%Y")
        end_date_input.send_keys(cu.get_date_string(end_date))
        time.sleep(1) 

        # Find and click the "Go" button
        go_button = wait_and_find_element(driver, By.ID, "btnRefreshGrid")
        go_button.click()
        time.sleep(5) 
        try:
            matching_records_element = driver.find_element(By.XPATH, "//td[@colspan='11' and contains(text(), 'No matching records found')]")
            raise Exception("No Signed Orders found")
        except Exception as e:
            pass

        destination_package = openpyxl.load_workbook(order_template)
        worksheet = destination_package.active
        first_page="1"
        while True:
            table_rows = driver.find_elements(By.CSS_SELECTOR, "#signed-docs-grid tbody tr")
            for row in table_rows:
                try:
                    document_id = cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(10) span.text-muted").text)
                    if is_document_to_be_filed(document_id,worksheet):
                        checkbox = driver.find_element(By.CSS_SELECTOR, 'input[value="'+document_id+'"]')
                        if checkbox.is_selected():
                            checkbox.click()
                        time.sleep(1)
                        checkbox.click()
                except Exception as e:
                    print(str(e))
                    continue
            try:
                button = driver.find_element(By.XPATH, "//*[@id='toolbar']/div/button[4]")
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(2)
                button.click()
                # driver.execute_script("document.querySelector('button.btn.btn-default:contains(\"File\")').click()")
                time.sleep(2)
                ok_button = driver.find_element(By.XPATH, '//*[@id="dialog"]/div/div/div[3]/button[2]')
                ok_button.click()
                time.sleep(10)
                break
            except Exception as ex:
                pass

                
            destination_package.save(order_template)
            if len(table_rows)<10:
                break
            next_button = driver.find_element(By.XPATH, "//li[@class='page-next']/a")
            try:
                next_button.click()
                time.sleep(5)
                current_page = driver.find_element(By.XPATH, "//li[@class='page-number active']/a").text
                if first_page == current_page:
                    break
            except Exception as e:
                break
        destination_package.save(order_template)
        destination_package.close()
        time.sleep(10)
        driver.quit() 
    except Exception as e:
        print(e)
        
# sendToPhysician("https://live.doctoralliance.com/Home/Login","rpaa723","rpa123","Axxess-StandardHomeHealth")