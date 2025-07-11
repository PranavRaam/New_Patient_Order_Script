import openpyxl
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
from datetime import datetime
import json
import time
import ReadConfig as rc
import os
import shutil
import CommonUtil as cu
import re

def upload_to_wavdb(reportFolderName):
    configuration = rc.readConfig()
    working_folder = cu.getFolderPath("O", reportFolderName)
    working_file=os.path.join(working_folder,"OrderTemplate.xlsx")
    api_key=configuration["APIKey"]
    api_url=configuration["APIBaseURL"]
    working_file = working_file.replace('\\', '/')
    try:
        upload_file(api_key,api_url,working_file)
    except Exception as e:
        raise Exception("Error in Map & WAV Upload Function: "+str(e))

def upload_file(api_key,api_url,working_file):
    api_endpoint=api_url+"api/Order/UploadBulkList"
    fileName=working_file.split('/')[-1]
    with open(working_file, 'rb') as file_stream:
        file_content = file_stream.read()

        multipart_encoder = MultipartEncoder(
            fields={'file': (fileName, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        )

        headers = {'Content-Type': multipart_encoder.content_type, 'X-SERVICE-KEY': api_key}

        try:
            response = requests.post(api_endpoint, data=multipart_encoder, headers=headers, timeout=300)
            response.raise_for_status()

            failed_list = get_failed_list_from_response(response)
            print(failed_list)
            update_excel_with_failure_comments(failed_list, working_file)

        except requests.exceptions.RequestException as e:
            print(f"Error: {e}")

def get_failed_list_from_response(response):
    failed_list = []
    try:
        failed_list_json = response.text
        failed_list = json.loads(failed_list_json)
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
    return failed_list

def update_excel_with_failure_comments(failed_list, new_file_path):
    package = openpyxl.load_workbook(new_file_path)
    worksheet = package.active
    try:
        for row in range(2, worksheet.max_row + 1):
            comment = next((f["value"] for f in failed_list if f["key"] == row), "")
            if "Created" in comment or "Updated" in comment:
                worksheet.cell(row, 16).value = comment.split("|")[0]
                worksheet.cell(row, 1).value = comment.split("|")[1]

            else:
                worksheet.cell(row, 16).value = "Failed: " + comment.split("|")[1]

        time.sleep(5)  # Wait for 10 seconds
        package.save(new_file_path)
    except Exception as e:
        package.save(new_file_path)
        raise str(e)