import openpyxl
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import json
import time
import ReadConfig as rc
import os
import shutil
import CommonUtil as cu
import pandas as pd
import xlrd
import pyautogui

def map_order(reportFolderName,credName,location):
    configuration = rc.readConfig()
    working_folder = cu.getFolderPath("O", reportFolderName)
    template_folder=configuration["OrderTemplatePath"]
    rpaName=configuration["RPA"]
    working_file=os.path.join(working_folder,"OrderTemplate.xlsx")
    for file_nameEl in os.listdir(working_folder):
        if file_nameEl.endswith(".xls") and file_nameEl.startswith("OrdersToBeSent_") :
            order_list_path=os.path.join(working_folder,file_nameEl)
    
    working_file = working_file.replace('\\', '/')
    order_list_path = order_list_path.replace('\\', '/')
    dfOrder = pd.read_excel(order_list_path, skiprows=1)
    shutil.copy2(template_folder, working_file)
    map_to_wav (working_file, order_list_path, dfOrder,rpaName,credName,location)


def map_to_wav(working_file, order_list_path, dfOrder,rpaName,credName,location):
    destination_package = openpyxl.load_workbook(working_file)
    destination_worksheet = destination_package.active
    dest_row = 2
    for index, row in dfOrder.iterrows():
        try:
            if pd.notna(row['Order Number']) and 'Total Number Of Orders' not in str(row['Order Number']):
                destination_worksheet.cell(dest_row, 2).value = cu.clean_null_data(row['Order Number'])
                destination_worksheet.cell(dest_row, 3).value = cu.clean_null_data(row['Patient'])
                destination_worksheet.cell(dest_row, 4).value = cu.get_date_string(cu.clean_null_data(row['Order Date']))
                destination_worksheet.cell(dest_row, 15).value = cu.clean_null_data(row['Type'])
                destination_worksheet.cell(dest_row, 17).value = cu.clean_null_data(rpaName)
                destination_worksheet.cell(dest_row, 18).value = cu.clean_null_data(credName)
                destination_worksheet.cell(dest_row, 19).value = cu.clean_null_data(location)

                dest_row += 1
        except Exception as e:
            destination_worksheet.cell(dest_row, 16).value= str(e)
            continue
    time.sleep(5)  # Wait for 10 seconds
    destination_package.save(working_file)




# map_and_upload("Axxess-StandardHomeHealth")
