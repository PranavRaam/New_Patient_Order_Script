import openpyxl
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
from datetime import datetime
import json
import time
import ReadConfig as rc
import os
import shutil
import CommonUtil as cu
import re

def upload_efax_to_wavdb(reportFolderName):
    configuration = rc.readConfig()
    working_folder = cu.getFolderPath("O", reportFolderName)
    working_file=os.path.join(working_folder,"EfaxTemplate.xlsx")
    api_key=configuration["APIKey"]
    api_url=configuration["APIBaseURL"]
    working_file = working_file.replace('\\', '/')
    destination_package = openpyxl.load_workbook(working_file)
    worksheet = destination_package.active
    for row in range(2, worksheet.max_row + 1):
        try:
            source= pgname=efaxno=efaxdate=hhah_name=patientName=docLabel=filePath= efax_wavid=""
            efax_wavid=cu.clean_null_data(worksheet.cell(row, 1).value)
            source=cu.clean_null_data(worksheet.cell(row, 2).value)
            pgname=cu.clean_null_data(worksheet.cell(row, 3).value)
            efaxno=cu.clean_null_data(worksheet.cell(row, 4).value)
            efaxdate=cu.clean_null_data(worksheet.cell(row, 5).value)
            hhah_name=cu.clean_null_data(worksheet.cell(row, 6).value)
            patientName=cu.clean_null_data(worksheet.cell(row, 7).value)
            docLabel=cu.clean_null_data(worksheet.cell(row, 15).value)
            efaxdocid=cu.clean_null_data(worksheet.cell(row, 20).value)
            filePath=get_filepath(api_key,api_url,working_folder,efaxdocid)
            if filePath and not efax_wavid:
                worksheet.cell(row, 17).value = filePath
                efax_wavid = upload_file(api_key,api_url, source, pgname,efaxno,efaxdate,hhah_name,patientName,docLabel,filePath,efaxdocid)
                worksheet.cell(row, 1).value=efax_wavid
                destination_package.save(working_file)

        except Exception as e:
            worksheet.cell(row, 19).value="Error in WAV Upload Function: "+str(e)

    destination_package.save(working_file)
    destination_package.close()


def upload_file(api_key,api_url, source, pgname,efaxno,efaxdate,hhah_name,patientName,docLabel,filePath,efaxdocid):
    api_endpoint=api_url+"api/Efax/PostEfaxData"
    headers = {'X-SERVICE-KEY': api_key}
    efax_id=""
    try:
        payload = {
            "source": source,
            "pgName": pgname,
            "ehR_EFAX_ID": efaxno,
            "efaX_Date": efaxdate,
            "hhaH_Name": hhah_name,
            "patientName": patientName,
            "dob": "",
            "hiClaim": "",
            "physicianNPI": "",
            "mrn": "",
            "episodeFrom": "",
            "episodeTo": "",
            "payorSource": "",
            "documentLabel": docLabel,
            "documentStatus": "",
            "efaxFilePath": filePath,
            "isEfaxDeleted": False,
            "ehR_EFAX_DOC_ID":efaxdocid
        }
        response = requests.post(api_endpoint, headers=headers, json=payload)
        if response.status_code == 200:
            efax_id = response.text
        
        return efax_id
    except Exception as ex:
        return str(ex)
    
def get_filepath(api_key,api_url,working_folder,efaxno):
    try:
        cloudPath=""
        api_endpoint = api_url + "api/Efax/UploadFileToBlob"
        file_path = os.path.join(working_folder,"EFax", efaxno+".pdf")
        file_path = file_path.replace('\\', '/')
        if os.path.exists(file_path):
            with open(file_path, 'rb') as file_stream:
                file_content = file_stream.read()

            multipart_encoder = MultipartEncoder(
                fields={'file': (efaxno+".pdf", file_content, 'application/pdf')}
            )
            headers = {
            'X-SERVICE-KEY': api_key,
            'Content-Type': multipart_encoder.content_type,
            }

            response = requests.post(api_endpoint, data=multipart_encoder, headers=headers, timeout=300)
            if response.status_code == 200:
                cloudPath=response.text

        return cloudPath
    except Exception as ex:
        return str(ex)


# upload_efax_to_wavdb("Athena-GAH")