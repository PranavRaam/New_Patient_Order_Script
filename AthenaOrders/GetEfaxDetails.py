import pdfplumber
import openpyxl
import json
import time
import ReadConfig as rc
import os
import re
import FetchAthenaConfig as fc
from datetime import datetime, timedelta
import CommonUtil as cu
import shutil
import requests
from urllib.parse import urlparse, parse_qs

def get_efax_doc_id(url):
    document_id=""
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    document_id = query_params.get('ID', [''])[0]
    return document_id

# Path to your PDF file
def get_efax_detail(credName, reportFolderName):
    configuration = rc.readConfig()
    rpaName = configuration["RPA"]
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/EfaxTemplate.xlsx"
    order_template = order_template.replace('\\', '/')
    download_folder=configuration["DownloadPath"]
    template_folder=configuration["EfaxTemplatePath"]
    shutil.copy2(template_folder, order_template)

    for signed in os.listdir(download_folder):
        if signed.endswith(".pdf"):
            file_path=os.path.join(download_folder,signed)
            os.remove(file_path)

    Efax_folder = os.path.join(working_folder, "EFax")
    Efax_folder = Efax_folder.replace('\\', '/')
    if os.path.exists(Efax_folder):
        shutil.rmtree(Efax_folder)

    os.makedirs(Efax_folder)

    Efax_Data_folder = os.path.join(working_folder, "EfaxData")
    Efax_Data_folder = Efax_Data_folder.replace('\\', '/')
    try:
        destination_package = openpyxl.load_workbook(order_template)
        worksheet = destination_package.active
        excel_row = worksheet.max_row+1
        efax_data_files = [os.path.join(Efax_Data_folder, file) for file in os.listdir(Efax_Data_folder) if os.path.isfile(os.path.join(Efax_Data_folder, file))]
        if len(efax_data_files) > 0:
            for efaxdata in efax_data_files:
                try:
                    efaxdata = efaxdata.replace('\\', '/')
                    text=""
                    links=[]
                    with pdfplumber.open(efaxdata) as pdf:
                        for page in pdf.pages:
                            text = page.extract_text()
                            # print(text)
                            annotations = page.annots
                            links = [annot['uri'] for annot in annotations if annot['uri'] ]
                            # for link in links:
                            #     print(link)
                            break
                    efax_list=[]
                    efax_list=extract_data(text, links)
                    for ef in efax_list:
                        for fax in ef['efax_detail']:
                            efaxdocid=get_efax_doc_id(fax['link'])
                            efaxUploaded = IsEfaxExists(efaxdocid)
                            if not efaxUploaded:
                                worksheet.cell(excel_row, 2).value = rpaName
                                worksheet.cell(excel_row, 3).value = credName
                                try:
                                    worksheet.cell(excel_row, 4).value = ef['efaxno']
                                except Exception as e:
                                    pass
                                try:
                                    worksheet.cell(excel_row, 5).value = ef['efaxdate']
                                except Exception as e:
                                    pass
                                try:
                                    worksheet.cell(excel_row, 7).value = fax['patientname']
                                except Exception as e:
                                    pass
                                try:
                                    worksheet.cell(excel_row, 15).value = fax['doclabel']
                                except Exception as e:
                                    pass
                                try:
                                    worksheet.cell(excel_row, 18).value = fax['link']
                                except Exception as e:
                                    pass
                                try:
                                    worksheet.cell(excel_row, 20).value = efaxdocid
                                except Exception as e:
                                    pass
                                excel_row=excel_row+1
                                destination_package.save(order_template)
                except Exception as e:
                    continue

        destination_package.save(order_template)
        destination_package.close()
    except Exception as e:
        logFolder = configuration["LogPath"]
        logFilepath=logFolder+"log_"+str(datetime.now().year)+"-"+str(datetime.now().month)+"-"+str(datetime.now().day)+".txt"
        cu.logFile(logFilepath, str(e))



def extract_data(text, links):
    try:
        efax_list = []
        current_efax = {}
        current_efax_detail=[]
        current_efax_links=[]
        efax_found=False
        for line in text.split('\n'):
            # Extract efax number
            efax_match = re.match(r'#(\d+)', line)
            if efax_match:
                efax_found=True
                if current_efax:
                    try:
                        if current_efax['efaxno']:
                            for k, link in enumerate(current_efax_links):
                                try:
                                    current_efax_detail[k]['link']=link
                                except Exception as e:
                                    if k>=len(current_efax_detail):
                                        efax_detail={}
                                        efax_detail['link']=link
                                        current_efax_detail.append(efax_detail)

                            current_efax['efax_detail']=current_efax_detail
                            efax_list.append(current_efax)
                    except Exception as e:
                        pass
                    current_efax = {}
                    current_efax_detail=[]
                    current_efax_links=[]
                current_efax_links=get_relevant_links(efax_match.group(1),links)
                current_efax = {'efaxno': efax_match.group(1)}
            
            if efax_found:
                # Extract efax date
                date_match = re.search(r'\d{1,2}/\d{1,2}/\d{4} \d{1,2}:\d{2}[ap]m', line)
                if date_match:
                    current_efax['efaxdate'] = date_match.group()
                
                # Handle 'Unknown' document label
                
                if 'Unknown' in line:
                    efax_detail={}
                    efax_detail['doclabel'] = 'Unknown'
                    efax_detail['patientname'] =""
                    efax_detail['link']=""
                    current_efax_detail.append(efax_detail)
                else:
                    # Extract patient name
                    efax_detail={}
                    patient_match = re.search(r',\s*([^,]+?),\s*([^,]+?)\s*\((\d+yo)?', line)
                    if patient_match:
                        last_name = patient_match.group(1).strip()
                        first_name = patient_match.group(2).strip()
                        full_name = f"{last_name}, {first_name}"
                        efax_detail['patientname']= full_name
                    
                    # Extract document label
                    doc_label_match = re.search(r'from \d{10} (.+?)(?:,|$)', line)
                    if doc_label_match:
                        efax_detail['doclabel'] = doc_label_match.group(1)
                    
                    if efax_detail:
                        efax_detail['link']=""
                        current_efax_detail.append(efax_detail)
        try:
            if current_efax and current_efax['efaxno']:
                for k, link in enumerate(current_efax_links):
                    try:
                        current_efax_detail[k]['link']=link
                    except Exception as e:
                        if k>=len(current_efax_detail):
                            efax_detail={}
                            efax_detail['link']=link
                            current_efax_detail.append(efax_detail)

                current_efax['efax_detail']=current_efax_detail
                efax_list.append(current_efax)
                
        except Exception as e:
            pass
        return efax_list
    except Exception as e:
        print(str(e))


def get_relevant_links(efax_no,links):
    e_links=[]
    for i, link in enumerate(links):
        if efax_no in link:
            for j in range(i + 1, len(links)):
                if 'SCANID' not in links[j]:
                    e_links.append(links[j])
                else:
                    break
    return e_links


def IsEfaxExists(efax_doc_no):
    efax_exists=False
    configuration={}
    configuration=rc.readConfig()
    api_key=configuration["APIKey"]
    api_url=configuration["APIBaseURL"]
    headers = {}
    headers["X-SERVICE-KEY"] = api_key
    endpoint = f"api/Efax/GetEfaxDataByEfaxDocId/{efax_doc_no}"
    api_endpoint = api_url + endpoint
    response = requests.get(api_endpoint, headers=headers)
    if response.status_code == 200:
        efax_exists=True
    return efax_exists

# get_efax_detail("Grace at Home", "Athena-GAH")