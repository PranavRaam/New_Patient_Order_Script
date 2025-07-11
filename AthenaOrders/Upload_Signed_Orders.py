from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
from openpyxl import load_workbook
import time
from datetime import datetime, timedelta
import os
import pandas as pd
import json
from datetime import datetime
import shutil
import ReadConfig as rc
import SplitPDF
import CommonUtil as cu
import openpyxl
def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def uploadSigned(url,email,password,credName,reportFolderName):
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/OrderTemplate.xlsx"
    order_template = order_template.replace('\\', '/')

    driver = webdriver.Chrome()
    driver.maximize_window()
    try:
        for ctr in range(3):
            try:
                cu.login_to_EHR(url,email,password,credName,driver)

                driver.switch_to.frame("GlobalNav")
                time.sleep(1)

                patients_menu = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='patientsmenucomponent']")))
                patients_menu.click()
                time.sleep(1)
                driver.switch_to.default_content()
                time.sleep(1)
                add_document_submenu = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@class="categoryitem" and text()="Add Document"]')))
                add_document_submenu.click()
                time.sleep(10)

                try:
                    driver.switch_to.frame("GlobalWrapper")
                    driver.switch_to.frame(0)
                    current_frame_attributes = driver.execute_script("return window.frameElement")

                    # Check if the frame has id attribute
                    if "id" in current_frame_attributes:
                        print("Current frame id:", current_frame_attributes["id"])
                        
                    # Check if the frame has name attribute
                    if "name" in current_frame_attributes:
                        print("Current frame name:", current_frame_attributes["name"])

                    for tag in driver.find_elements(By.TAG_NAME,'frame'):
                        print (tag.id)
                    driver.switch_to.frame(0)
                  
                    driver.switch_to.frame("frMain")
                    
                    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@class="name accordion-trigger clickable" and text()="Upload files from computer or network"]')))
                    element.click()
                    time.sleep(5)
                except Exception as e:
                    pass

                destination_package = openpyxl.load_workbook(order_template)
                worksheet = destination_package.active
                end_row = worksheet.max_row
                for row in range(2, end_row + 1):
                    upload_status= cu.clean_null_data(worksheet.cell(row, 13).value)
                    signed_date=cu.get_date_string(cu.clean_null_data(worksheet.cell(row, 9).value))
                    docId=cu.get_date_string(cu.clean_null_data(worksheet.cell(row, 14).value))
                    doc_path=os.path.join(working_folder,"SignedOrders")
                    doc_path=os.path.join(doc_path,docId+".pdf")
                    doc_path = doc_path.replace('\\', '/')

                    #Status Changed
                    if signed_date and not upload_status:
                        try:
                            dropdown = Select(driver.find_element(By.ID, "add-document-class-select"))
                            dropdown.select_by_value("ADMIN")
                            time.sleep(2)  
                            file_input = driver.find_element_by_id("filedata")
                            file_input.send_keys(doc_path)
                            time.sleep(2)  
                            button = driver.find_element(By.CLASS_NAME, "add-document-create-button")
                            button.click()
                            time.sleep(30)  
                        except Exception as ex:
                            worksheet.cell(row, 20).value = str(ex)
                    try:
                        close_link = driver.find_element_by_link_text("Close X")
                        close_link.click()
                        time.sleep(2) 
                    except Exception as e:
                        pass

                destination_package.save(order_template)
                driver.quit()
            except Exception as ex:
                print(str(ex))
    except Exception as ex:
                print(str(ex))
        
# sendSigned("https://accounts.axxessweb.com/Login","cliffard@doctoralliance.com","Dallas@1234","Standards Home Care","Axxess-StandardHomeHealth")