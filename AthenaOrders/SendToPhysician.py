import openpyxl
import json
import time
import ReadConfig as rc
import os
import re
import FetchAthenaConfig as fc
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import calendar
import pyautogui
import CommonUtil as cu
import requests

def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def click_patient_select_element(driver):
    patient_select_element = wait_and_find_element(driver, By.CSS_SELECTOR, "span.select2-selection__rendered span.select2-selection__placeholder")
    patient_select_element.click()
    time.sleep(2)

def interact_with_search_field(driver, search_term):
    input_css_selector = "input.select2-search__field"
    search_input = wait_and_find_element(driver, By.CSS_SELECTOR, input_css_selector)
    search_input.clear()
    search_input.send_keys(search_term)
    time.sleep(10)


def sendToPhysician(da_url, da_login, da_password, reportFolderName):
    try:
        configuration={}
        configuration=rc.readConfig()
        efaxFolderPath=configuration["OrderFolderPath"]
        currDateStr=datetime.now().strftime("%Y-%m-%d") 
        working_folder = os.path.join(efaxFolderPath, currDateStr)
        working_folder = working_folder.replace('\\', '/')

        efax_folder = os.path.join(working_folder, reportFolderName)
        efax_folder = efax_folder.replace('\\', '/')
        
        efax_file = working_folder + "/EfaxTemplate.xlsx"

        destination_package = openpyxl.load_workbook(efax_file)
        worksheet = destination_package.active
        end_row = worksheet.max_row
        driver = webdriver.Chrome()
        driver.maximize_window()
        try:
            cu.login_to_da(da_url, da_login, da_password, driver)

            navbar_toggle = wait_and_find_element(driver,By.CSS_SELECTOR, "a.navbar-minimalize.btn.btn-primary")
            navbar_toggle.click()
            time.sleep(2)

            for row in range(2, end_row + 1):
                try:
                    id = str(worksheet.cell(row, 1).value)
                    filename = id + ".pdf"
                    patientname = cu.clean_null_data(worksheet.cell(row, 7).value)
                    efax_path = os.path.join(efax_folder, filename)
                    efax_path = efax_path.replace('\\', '/')

                    if not os.path.exists(efax_path):
                        worksheet.cell(row, 19).value = "Efax file not found"
                        continue

                    verbalOrderLink = wait_and_find_element(driver,By.CSS_SELECTOR, "a[href*='CreateSignable?docType=ORDER'] i.fa.fa-clipboard.fa-lg.fa-fw")
                    verbalOrderLink.click()
                    time.sleep(2)

                    uploadLink = wait_and_find_element(driver,By.CSS_SELECTOR, "a[href*='Upload?docType=ORDER']")
                    uploadLink.click()
                    time.sleep(2)

                    click_patient_select_element(driver)
                    interact_with_search_field(driver, patientname)

                    no_results_element = wait_and_find_element(driver,By.CLASS_NAME, "select2-results__option")
                    if "No results found" in no_results_element.text:
                        upload_doc_status(id, "Failed - Patient doesn't exist")
                        worksheet.cell(row, 19).value = "No Patient found in DA"
                        continue 

                    first_row_element = driver.find_element(By.XPATH, "//ul[@id='select2-PatientSelect2-results']/li[1]")
                    first_row_element.click()
                    time.sleep(2)

                    try:
                        file_input = wait_and_find_element(driver, By.XPATH, "//input[@type='file' and @name='docfile[]' and @id='docfile' and @accept='.pdf' and @class='form-control']")
                        file_input.send_keys(efax_path)
                        time.sleep(5)

                    except Exception as e:
                        upload_doc_status(id, "Failed to Upload")
                        worksheet.cell(row, 19).value = "Error occurred while uploading file"
                        continue

                    click_element(wait_and_find_element(driver, By.CLASS_NAME, 'btn.btn-default.fileinput-upload.fileinput-upload-button'))
                    time.sleep(5)

                    click_element(wait_and_find_element(driver, By.ID, 'SendPhysicianButton'))
                    time.sleep(5)

                    # Find the first row in the table
                    first_row = wait_and_find_element(driver, By.CSS_SELECTOR, "#unsigned-docs-grid tbody tr:first-child")

                    # Extract doc_id from the first row
                    doc_id_cell = first_row.find_element(By.CSS_SELECTOR, "td:nth-child(9)")
                    doc_id = doc_id_cell.text.strip()

                    worksheet.cell(row, 19).value = str(doc_id)

                    destination_package.save(efax_file)


                except Exception as e:
                    upload_doc_status(id, "Failed to Upload")
                    worksheet.cell(row, 19).value = str(e)
                    continue

            time.sleep(10)
            driver.quit()
            print("Worked dont worry")
            destination_package.save(efax_file)
        except Exception as e:
            print(f"Oh Exception : {e}")
    except Exception as e:
        raise e
    

def upload_doc_status(id, message):
    configuration={}
    configuration=rc.readConfig()
    api_key=configuration["APIKey"]
    api_url=configuration["APIBaseURL"]
    headers = {}
    headers["X-SERVICE-KEY"] = api_key
    endpoint = "api/Efax/UpdateDocUploadStatus/" + id + "/" + message
    api_endpoint = api_url + endpoint
    requests.put(api_endpoint, headers=headers)
        