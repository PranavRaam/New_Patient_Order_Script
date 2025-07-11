import openpyxl
from datetime import datetime,timedelta
import json
import time
import ReadConfig as rc
import os
import shutil
import CommonUtil as cu
import pandas as pd

def Check_DA_Creation(cred):
    try:
        current_working_folder= cu.getFolderPath("P",cred)
        previous_date_working_folder=cu.getPrevDayWorkingFolder("P",cred)
        curr_date_str = datetime.now().strftime("%Y-%m-%d")
        prev_date_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        working_file=current_working_folder+"/AgencyTemplate_"+curr_date_str+".xlsx"
        prev_working_file=previous_date_working_folder+"/AgencyTemplate_"+prev_date_str+".xlsx"

        if (os.path.exists(prev_working_file)):
            df = pd.read_excel(prev_working_file)
            package = openpyxl.load_workbook(working_file)
            worksheet = package.active
            for row in range(2, worksheet.max_row + 1):
                mrn=cu.clean_null_data(str(worksheet.cell(row, 18).value))
                if mrn:
                    if ((df['Medical Record No'].astype(str).str.strip() == mrn) & (df['DA Upload Status'] == "Passed")).any():
                        worksheet.cell(row, 68).value="Passed"
                        worksheet.cell(row, 69).value="Already Exists"
            package.save(working_file)
    except Exception as e:
        raise e


# Check_DA_Creation("Standard Home Health")




