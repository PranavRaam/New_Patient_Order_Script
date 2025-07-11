import openpyxl
import json
import time
import ReadConfig as rc
import os
import re
import FetchAthenaConfig as fc
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import calendar
import pyautogui
import CommonUtil as cu
import shutil
import pdfplumber


def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def signedOrderExtraction(reportFolderName):
    base_folder = cu.getFolderPath("O", reportFolderName)
    subfolders = [f.path for f in os.scandir(base_folder) if f.is_dir()]
    for fol in subfolders:
        if "SignedOrders" in fol:
            signed_orders_folder=fol
            break
    working_file=os.path.join(base_folder,"OrderTemplate.xlsx")
    working_file = working_file.replace('\\', '/')
    signed_orders_folder = signed_orders_folder.replace('\\', '/')

    # if signed_orders_folder:
    #     # Get a list of PDF files in the BulkOrders folder
    #     pdf_files = [f.path for f in os.scandir(signed_orders_folder) if f.is_file() and f.name.lower().endswith('.pdf')]
    try:
        destination_package = openpyxl.load_workbook(working_file)
        destination_worksheet = destination_package.active
        for row in range(2, destination_worksheet.max_row + 1):
            try:
                doc_id=cu.clean_null_data(destination_worksheet.cell(row, 14).value)
                fetched_order_no=cu.clean_null_data(destination_worksheet.cell(row, 2).value)
                signed_by_physician = cu.clean_null_data(destination_worksheet.cell(row, 12).value)
                if signed_by_physician and not fetched_order_no:
                    signed_order_files = [os.path.join(signed_orders_folder, file) for file in os.listdir(signed_orders_folder) if os.path.isfile(os.path.join(signed_orders_folder, file))]
                    start_date,end_date,mrn,npi,orderno,signed_date = get_order_detail(signed_order_files, doc_id)
                    destination_worksheet.cell(row, 7).value=mrn
                    if start_date:
                        destination_worksheet.cell(row, 5).value=cu.get_date_string(start_date)
                    if end_date:
                        destination_worksheet.cell(row, 6).value=cu.get_date_string(end_date)
                    # destination_worksheet.cell(row, 23).value=str(npi)
                    destination_worksheet.cell(row, 2).value=str(orderno)

            except Exception as e:
                destination_worksheet.cell(row, 20).value=str(e)

        destination_package.save(working_file)
    except Exception as e:
        print(e)


def get_order_detail(pdf_files,doc_id):
    start_date=""
    end_date=""
    mrn=""
    npi=""
    orderno=""
    signed_date=""
    for pdf in pdf_files:
        pdf_order = pdf.replace('\\', '/')
        if str(doc_id) in pdf_order:
            start_date,end_date,mrn,npi,orderno,signed_date=extract_pdf(pdf_order)
            break
    return start_date,end_date,mrn,npi,orderno,signed_date


def extract_pdf(pdf_order):
    try:
        start_date=""
        end_date=""
        mrn=""
        npi=""
        orderno=""
        signed_date=""
        with pdfplumber.open(pdf_order) as pdf:
            for page in pdf.pages:
                if mrn and start_date and end_date:
                    break
                text = page.extract_text()
                # text=text.replace('\n','')
                print(text)
                if not text:
                    raise Exception("PDF is not readable!")
             
                patterns = [r'Order#: (\d+)', r'Order: (\d+)', r'Order: #(\d+)', r'Order #(\d+)', r'Order# (\d+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        orderno = match.group(1) 
                        break
                
                patterns = [r'MR#: (\d+)', r'MRN#: (\d+)', r'MRN: (\d+)', r'MR: (\d+)', r'MRN #(\d+)', r'MR #(\d+)', r'MR: (\w+)', r'MRN: (\w+)', r'MR:(\w+)', r'MRN:(\w+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        mrn = match.group(1) 
                        break

                patterns = [r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4})', r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4})']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        start_date = match.group(1)
                        end_date = match.group(2)
                        break
                                
                if not mrn and not start_date and not end_date:
                    pattern = r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                    match = re.search(pattern, text)
                    if match:
                        start_date = match.group(1)
                        end_date = match.group(2)
                        mrn = match.group(3)
                        break
                    else:
                        pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                        match = re.search(pattern, text)
                        if match:
                            start_date = match.group(1)
                            end_date = match.group(2)
                            mrn = match.group(3)
                            break
                        else:
                            pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+-\d+)'
                            match = re.search(pattern, text)
                            if match:
                                start_date = match.group(1)
                                end_date = match.group(2)
                                mrn = match.group(3)
                                break
                            else:
                                pattern = r'Episode Start Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    start_date = match.group(1)
                                pattern = r'Episode End Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    end_date = match.group(1)
                                    break
                
        return start_date,end_date,mrn,npi,orderno,signed_date
    except Exception as e:
        raise e


        
# sendToPhysician("https://live.doctoralliance.com/Home/Login","rpaa723","rpa123","Axxess-StandardHomeHealth")