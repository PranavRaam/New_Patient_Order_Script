import requests
import ReadConfig as rc
from datetime import datetime
import CommonUtil as cu
import openpyxl
import json

def execute_da_patient_creation(daToken):
    daAPITokenUrl,daAPIPatientUrl,daAPITokenUserName,daAPITokenPswd,daAPITokenClinicianId,daAPITokenCaretakerId,reportFolderName=daToken
    try:
        current_date = datetime.now()
        formatted_date = current_date.strftime("%Y-%m-%d")
        current_working_folder= cu.getFolderPath("O",reportFolderName)
        working_file=current_working_folder+"/OrderTemplate_"+formatted_date+".xlsx"
        # working_file="S:\\DA\\DA\\AxxessPatientPython\\AgencyTemplate_2024-01-29.xlsx" #hardcoded to test
        working_file = working_file.replace('\\', '/')
        access_token = get_access_token(daAPITokenUrl,daAPITokenUserName,daAPITokenPswd)
        source_package = openpyxl.load_workbook(working_file)
        source_worksheet = source_package.active
        end_row = source_worksheet.max_row
        for row in range(2, end_row + 1):
            try:
                da_Patient_Exists=cu.clean_null_data(source_worksheet.cell(row, 69).value) 
                if da_Patient_Exists not in ["Already Exists","Created"]:

                    full_name = cu.clean_null_data(source_worksheet.cell(row, 3).value) 
                    name_parts= full_name.split(", ")
                    if len(name_parts) == 2:
                        ld, fd = name_parts
                        md = None
                    elif len(name_parts) == 3:
                        ld, fd, md = name_parts
                    firstName = cu.clean_null_data(fd)
                    middleName = cu.clean_null_data(md)
                    lastName = cu.clean_null_data(ld)

                    dob = cu.clean_null_data(source_worksheet.cell(row, 6).value) 
                    
                    payor = cu.clean_null_data(source_worksheet.cell(row, 28).value) 
                    serviceLine = cu.clean_null_data(source_worksheet.cell(row, 19).value) 
                    certFrom = cu.clean_null_data(source_worksheet.cell(row, 13).value) 
                    certTo = cu.get_episode_end_date(cu.clean_null_data(source_worksheet.cell(row, 14).value),certFrom,serviceLine)
                    hiClaim = cu.clean_null_data(source_worksheet.cell(row, 55).value) 
                    physicianNPI = cu.clean_null_data(source_worksheet.cell(row, 43).value) 
                    patientSex = cu.clean_null_data(source_worksheet.cell(row, 8).value) 
                    mrn = cu.clean_null_data(source_worksheet.cell(row, 7).value) #MRN
                    errComment=""
                    isValid,errComment=cu.isValidData(firstName,lastName,dob,certFrom,certTo,physicianNPI)
                    if isValid:
                        comment,status = create_patient_in_da(access_token,daAPIPatientUrl,firstName,middleName,lastName,dob,payor,certFrom,certTo,hiClaim,physicianNPI,patientSex,mrn,daAPITokenClinicianId,daAPITokenCaretakerId)
                        source_worksheet.cell(row, 68).value=status
                        source_worksheet.cell(row, 69).value=comment
                    else:
                        source_worksheet.cell(row, 68).value="Failed"
                        source_worksheet.cell(row, 69).value=errComment
            except Exception as e:
                source_worksheet.cell(row, 68).value="Failed"
                source_worksheet.cell(row, 69).value=str(e)
                
        source_package.save(working_file)
    except Exception as e:
        raise Exception("Error in Create DA Patient: "+str(e))


def get_access_token(daAPITokenUrl,daAPITokenUserName,daAPITokenPswd):
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    data = {
        'grant_type': 'password',
        'username': daAPITokenUserName,
        'password': daAPITokenPswd
    }
    response = requests.post(daAPITokenUrl, headers=headers, data=data)
    if response.status_code == 200:
        json_response = response.json()
        access_token = json_response.get('access_token')
        return access_token
    else:
        raise Exception(f"Error: {response.status_code} - {response.text}")
    

def create_patient_in_da(access_token,daAPIPatientUrl,firstName,middleName,lastName,dob,payor,certFrom,certTo,hiClaim,physicianNPI,patientSex,mrn,daAPITokenClinicianId,daAPITokenCaretakerId):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + access_token
    }
    fullname= firstName + " "+ middleName+ " "+ lastName if middleName else firstName + " "+ lastName
    if not hiClaim:
        hiClaim="0000"
    data = {
    "patientInfo": {
        "name": fullname,
        "fullAddress": None,
        "id": 0,
        "firstName": firstName,
        "middleInitial": middleName,
        "lastName": lastName,
        "sex": patientSex,
        "dob": dob,
        "suite": None,
        "address": None,
        "phone": None,
        "state": None,
        "city": None,
        "zipCode": None,
        "paySource": payor,
        "insuranceNumber": hiClaim,
        "medicalRecordNumber": mrn,
        "ssn": None,
        "physicianHelperId": 0,
        "externalId": None,
        "status": None,
        "physicianName": "",
        "physicianNpi": physicianNPI
    },
    "patientStatus": {
        "medications": None,
        "dmeAndSupplies": None,
        "safetyMeasures": None,
        "nutritionalRequirements": None,
        "allergies": None,
        "functionalLimitations": {
        "amputation": True,
        "bowel": True,
        "contracture": True,
        "hearing": True,
        "paralysis": True,
        "endurance": True,
        "ambulation": True,
        "speech": True,
        "legallyBlind": True,
        "dyspnea": True,
        "others": None
        },
        "activitiesPermitted": {
        "completeBedrest": True,
        "bedrestBRP": True,
        "upAsTolerated": True,
        "transferBedChair": True,
        "exercisesPrescribed": True,
        "partialWeightBearing": True,
        "independentAtHome": True,
        "crutches": True,
        "cane": True,
        "wheelChair": True,
        "walker": True,
        "noRestrictions": True,
        "others": None
        },
        "mentalStatus": {
        "oriented": True,
        "forgetful": True,
        "comatose": True,
        "depressed": True,
        "disoriented": True,
        "lethargic": True,
        "agitated": True,
        "others": None
        },
        "prognosis": "poor",
        "orders": None,
        "goals": None,
        "state": "admitted",
        "startOfCareDate": certFrom,
        "certPeriodFrom": certFrom,
        "certPeriodTo": certTo,
        "diagnoses": [
        {
            "description": "string",
            "code": "string",
            "eorO": "1",
            "diagnosisType": "principal",
            "date": "2024-01-24T02:43:14.255Z"
        }
        ]
    },
    "physicianNpi": physicianNPI,
    "clinicianId": {
        "npi": "NULL",
        "id": daAPITokenClinicianId,
        "externalId": "NULL"
    },
    "careProviderId": {
        "npi": "NULL",
        "id": daAPITokenCaretakerId,
        "externalId": "NULL"
    }
    }

    response = requests.post(daAPIPatientUrl, headers=headers, data=json.dumps(data))
    comment=""
    status=""
    response_json = json.loads(response.content)
    if response_json['isSuccess']:
        status="Passed"
        action_type = response_json['value']['actionType']
        id = response_json['value']['id']
        comment=action_type +":" + str(id)
    else:
        status="Failed"
        error_message = response_json['errorMessage']
        comment=error_message

    return comment,status

    
# execute_da_patient_creation("Axxess-StandardHomeHealth")