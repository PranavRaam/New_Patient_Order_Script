import fitz  # PyMuPDF
import re
import os
import CommonUtil as cu
import openpyxl
import pandas as pd
import pdfplumber


def bulkOrderOpening(reportFolderName):
    base_folder = cu.getFolderPath("O", reportFolderName)
    subfolders = [f.path for f in os.scandir(base_folder) if f.is_dir()]
    bulk_orders_folder=subfolders[0]
    working_file=os.path.join(base_folder,"OrderTemplate.xlsx")
    working_file = working_file.replace('\\', '/')
    if bulk_orders_folder:
        # Get a list of PDF files in the BulkOrders folder
        pdf_files = [f.path for f in os.scandir(bulk_orders_folder) if f.is_file() and f.name.lower().endswith('.pdf')]
        destination_package = openpyxl.load_workbook(working_file)
        destination_worksheet = destination_package.active
        for row in range(2, destination_worksheet.max_row + 1):
            try:
                order_no=destination_worksheet.cell(row, 2).value
                patient_name=destination_worksheet.cell(row, 3).value
                document_name=destination_worksheet.cell(row, 15).value
                start_date,end_date,mrn,npi,DOB = get_order_detail(pdf_files,order_no,patient_name,document_name)
                destination_worksheet.cell(row, 7).value=mrn
                destination_worksheet.cell(row, 5).value=cu.get_date_string(start_date)
                destination_worksheet.cell(row, 6).value=cu.get_date_string(end_date)
                destination_worksheet.cell(row, 22).value=cu.get_date_string(DOB)
                destination_worksheet.cell(row, 23).value=str(npi)
            except Exception as e:
                destination_worksheet.cell(row, 20).value=str(e)

        destination_package.save(working_file)
    else:
        print(f"BulkOrders folder not found in {base_folder}")


def get_order_detail(pdf_files,order_no,patient_name,document_name):
    start_date=""
    end_date=""
    mrn=""
    npi=""
    DOB=""
    for pdf in pdf_files:
        pdf_order = pdf.replace('\\', '/')
        if str(order_no) in pdf_order:
            start_date,end_date,mrn,npi,DOB=extract_pdf(pdf_order)
            break
    if not mrn:
        for pdf in pdf_files:
            if mrn:
                break
            pdf_order = pdf.replace('\\', '/')
            if "Order_" in pdf_order:
                start_date,end_date,mrn,npi,DOB=extract_and_rename_pdf(pdf_order,order_no,patient_name,document_name)
    return start_date,end_date,mrn,npi,DOB


def extract_pdf(pdf_order):
    try:
        start_date=""
        end_date=""
        mrn=""
        npi=""
        DOB=""
        with pdfplumber.open(pdf_order) as pdf:
            for page in pdf.pages:
                if mrn and start_date and end_date:
                    break
                text = page.extract_text()
                # text=text.replace('\n','')
                print(text)
                patterns = [r'NPI: (\d+)', r'NPI #(\d+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        npi = match.group(1) 
                        break
                pattern = r'DOB: (\d{1,2}/\d{1,2}/\d{4})'
                match = re.search(pattern, text)
                if match:
                    DOB = match.group(1)
                patterns = [r'MR#: (\d+)', r'MRN#: (\d+)', r'MRN: (\d+)', r'MR: (\d+)', r'MRN #(\d+)', r'MR #(\d+)', r'MR: (\w+)', r'MRN: (\w+)', r'MR:(\w+)', r'MRN:(\w+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        mrn = match.group(1) 
                        if start_date=="":
                            patternDate = r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4})'
                            match = re.search(patternDate, text)
                            if match:
                                start_date = match.group(1)
                                end_date = match.group(2)
                                break 
                            else:
                                patternDate = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(patternDate, text)
                                if match:
                                    start_date = match.group(1)
                                    end_date = match.group(2)
                                    break 
                                else:
                                    pattern = r'Episode Start Date: (\d{1,2}/\d{1,2}/\d{4})'
                                    match = re.search(pattern, text)
                                    if match:
                                        start_date = match.group(1)
                                    pattern = r'Episode End Date: (\d{1,2}/\d{1,2}/\d{4})'
                                    match = re.search(pattern, text)
                                    if match:
                                        end_date = match.group(1)
                                        break
                                
                if not mrn:
                    pattern = r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                    match = re.search(pattern, text)
                    if match:
                        start_date = match.group(1)
                        end_date = match.group(2)
                        mrn = match.group(3)
                        break
                    else:
                        pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                        match = re.search(pattern, text)
                        if match:
                            start_date = match.group(1)
                            end_date = match.group(2)
                            mrn = match.group(3)
                            break
                        else:
                            pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+-\d+)'
                            match = re.search(pattern, text)
                            if match:
                                start_date = match.group(1)
                                end_date = match.group(2)
                                mrn = match.group(3)
                                break
                            else:
                                pattern = r'Episode Start Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    start_date = match.group(1)
                                pattern = r'Episode End Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    end_date = match.group(1)
                                    break
                
        return start_date,end_date,mrn,npi,DOB
    except Exception as e:
        raise e


def extract_and_rename_pdf(pdf_order,order_no,patient_name,document_name):
    start_date=""
    end_date=""
    mrn=""
    npi=""
    DOB=""
    patientFName=patient_name.split(',')[1].strip()
    with pdfplumber.open(pdf_order) as pdf:
        for page in pdf.pages:
            if mrn and start_date and end_date:
                break
            text = page.extract_text()
            # print(text)
            if patientFName in text and "Auto-Generated" in document_name and "Auto-Generated" in text:
                patterns = [r'NPI: (\d+)', r'NPI #(\d+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        npi = match.group(1) 
                        break
                pattern = r'DOB: (\d{1,2}/\d{1,2}/\d{4})'
                match = re.search(pattern, text)
                if match:
                    DOB = match.group(1)
                patterns = [r'MR#: (\d+)', r'MRN#: (\d+)', r'MRN: (\d+)', r'MR: (\d+)', r'MRN #(\d+)', r'MR #(\d+)', r'MR: (\w+)', r'MRN: (\w+)', r'MR:(\w+)', r'MRN:(\w+)']
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        mrn = match.group(1) 
                        if start_date=="":
                            patternDate = r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4})'
                            match = re.search(patternDate, text)
                            if match:
                                start_date = match.group(1)
                                end_date = match.group(2)
                                break 
                            else:
                                patternDate = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(patternDate, text)
                                if match:
                                    start_date = match.group(1)
                                    end_date = match.group(2)
                                    break 
                                else:
                                    pattern = r'Episode Start Date: (\d{1,2}/\d{1,2}/\d{4})'
                                    match = re.search(pattern, text)
                                    if match:
                                        start_date = match.group(1)
                                    pattern = r'Episode End Date: (\d{1,2}/\d{1,2}/\d{4})'
                                    match = re.search(pattern, text)
                                    if match:
                                        end_date = match.group(1)
                                    break
                if not mrn:
                    pattern = r'(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                    match = re.search(pattern, text)
                    if match:
                        start_date = match.group(1)
                        end_date = match.group(2)
                        mrn = match.group(3)
                        break
                    else:
                        pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+)'
                        match = re.search(pattern, text)
                        if match:
                            start_date = match.group(1)
                            end_date = match.group(2)
                            mrn = match.group(3)
                            break
                        else:
                            pattern = r'(\d{1,2}/\d{1,2}/\d{4})-(\d{1,2}/\d{1,2}/\d{4}) (\d+-\d+)'
                            match = re.search(pattern, text)
                            if match:
                                start_date = match.group(1)
                                end_date = match.group(2)
                                mrn = match.group(3)
                                break
                            else:
                                pattern = r'Episode Start Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    start_date = match.group(1)
                                pattern = r'Episode End Date: (\d{1,2}/\d{1,2}/\d{4})'
                                match = re.search(pattern, text)
                                if match:
                                    end_date = match.group(1)
                                break
    directory, old_name = os.path.split(pdf_order)
    new_file_path = os.path.join(directory, str(order_no)+".pdf")
    os.rename(pdf_order, new_file_path)
    return start_date,end_date,mrn,npi,DOB





