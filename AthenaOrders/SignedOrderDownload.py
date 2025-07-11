import openpyxl
import json
import time
import ReadConfig as rc
import os
import re
import FetchAthenaConfig as fc
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import calendar
import pyautogui
import CommonUtil as cu
import shutil

def wait_and_find_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )

def click_element(element):
    try:
        element.click()
    except Exception as ex:
        print(f"Error clicking element: {str(ex)}")
        exit()

def hover_and_click(driver, script, element):
    try:
        driver.execute_script(script, element)
        time.sleep(2)
    except Exception as ex:
        print(f"Error executing script: {str(ex)}")
        exit()

def click_patient_select_element(driver):
    patient_select_element = wait_and_find_element(driver, By.CSS_SELECTOR, "span.select2-selection__rendered span.select2-selection__placeholder")
    patient_select_element.click()
    time.sleep(2)

def interact_with_search_field(driver, search_term):
    input_css_selector = "input.select2-search__field"
    search_input = wait_and_find_element(driver, By.CSS_SELECTOR, input_css_selector)
    search_input.clear()
    search_input.send_keys(search_term)
    time.sleep(10)


def download_signed_orders(da_url, da_login, da_password, reportFolderName, location, credName, helper_id):

    configuration = rc.readConfig()
    rpaName = configuration["RPA"]
    working_folder = cu.getFolderPath("O", reportFolderName)
    order_template = working_folder + "/OrderTemplate.xlsx"
    order_template = order_template.replace('\\', '/')
    download_folder=configuration["DownloadPath"]
    template_folder=configuration["OrderTemplatePath"]
    
    print(f"Setting up template from: {template_folder}")
    shutil.copy2(template_folder, order_template)

    for signed in os.listdir(download_folder):
        if signed.endswith(".pdf"):
            file_path=os.path.join(download_folder,signed)
            os.remove(file_path)

    signed_orders_folder = os.path.join(working_folder, "SignedOrders")
    signed_orders_folder = signed_orders_folder.replace('\\', '/')
    if os.path.exists(signed_orders_folder):
        shutil.rmtree(signed_orders_folder)

    os.makedirs(signed_orders_folder)

    print("Starting browser automation...")
    try:
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        
        # Simplified Chrome configuration based on working doctoralliance_bot.py
        chrome_options = Options()
        # Remove headless mode to see the browser
        # chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        print("Initializing Chrome driver...")
        try:
            # Try the newer ChromeDriver version first
            service = Service("/home/lone/.cache/selenium/chromedriver/linux64/137.0.7151.55/chromedriver")
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.maximize_window()
            print("Chrome driver initialized successfully with ChromeDriver 137")
        except Exception as e:
            print(f"Error with ChromeDriver 137: {e}")
            print("Trying with webdriver-manager...")
            try:
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.maximize_window()
                print("Chrome driver initialized successfully with webdriver-manager")
            except Exception as e2:
                print(f"Error with webdriver-manager: {e2}")
                print("Trying Firefox as fallback...")
                try:
                    from selenium.webdriver.firefox.options import Options as FirefoxOptions
                    from selenium.webdriver.firefox.service import Service as FirefoxService
                    from webdriver_manager.firefox import GeckoDriverManager
                    
                    firefox_options = FirefoxOptions()
                    # Remove headless mode to see the browser
                    # firefox_options.add_argument("--headless")
                    
                    firefox_service = FirefoxService(GeckoDriverManager().install())
                    driver = webdriver.Firefox(service=firefox_service, options=firefox_options)
                    driver.maximize_window()
                    print("Firefox driver initialized successfully as fallback")
                except Exception as e3:
                    print(f"All browser initialization attempts failed. ChromeDriver 137 error: {e}, webdriver-manager error: {e2}, Firefox error: {e3}")
                    raise Exception("Could not initialize any web browser")
        
        try:
            print(f"Logging into DA: {da_url}")
            cu.login_to_da(da_url, da_login, da_password, driver)
            print("Login successful")
            
            print("Navigating to search page...")
            driver.get("https://backoffice.doctoralliance.com/Search")
            time.sleep(7)
            query_input = driver.find_element(By.ID, "Query")
            query_input.send_keys(helper_id)
            time.sleep(2)
            search_type_dropdown = driver.find_element(By.ID, "select2-SearchType-container")
            search_type_dropdown.click()
            time.sleep(1)
            input_field = driver.find_element(By.CLASS_NAME, "select2-search__field")
            input_field.send_keys("Users")
            time.sleep(3)
            wait = WebDriverWait(driver,10)
            first_result = wait.until(EC.visibility_of_element_located((By.XPATH, "//li[contains(@id, 'select2-SearchType-result')][1]")))
            first_result.click()
            time.sleep(3)
            button = driver.find_element(By.CLASS_NAME, "btn-success")
            button.click()
            time.sleep(5)
            row = driver.find_element(By.CLASS_NAME, "linkedRow")
            row.click()
            time.sleep(5)

            link = driver.find_element(By.LINK_TEXT,"Impersonate")
            link.click()
            time.sleep(7)

            driver.switch_to.window(driver.window_handles[1])

            signed_link = wait_and_find_element(driver, By.XPATH, "//a[contains(@href, '/Documents/Signed')]")
            signed_link.click()
            time.sleep(2) 

            start_date_input = wait_and_find_element(driver, By.ID, "StartDatePicker")
            start_date_input.clear()
            start_date=(datetime.now() - timedelta(days=5)).strftime("%m/%d/%Y")
            time.sleep(5)
            start_date_input.send_keys(cu.get_date_string(start_date))
            time.sleep(1) 

            # Find the end date input field and fill it with the current date
            end_date_input = wait_and_find_element(driver, By.ID, "EndDatePicker")
            end_date_input.clear()
            end_date=datetime.now().strftime("%m/%d/%Y")
            end_date_input.send_keys(cu.get_date_string(end_date))
            time.sleep(1) 

            # Find and click the "Go" button
            go_button = wait_and_find_element(driver, By.ID, "btnRefreshGrid")
            go_button.click()
            time.sleep(5) 
            try:
                matching_records_element = driver.find_element(By.XPATH, "//td[@colspan='11' and contains(text(), 'No matching records found')]")
                raise Exception("No Signed Orders found")
            except Exception as e:
                pass

            destination_package = openpyxl.load_workbook(order_template)
            worksheet = destination_package.active
            excel_row = worksheet.max_row+1
            first_page="1"
            while True:
                table_rows = driver.find_elements(By.CSS_SELECTOR, "#signed-docs-grid tbody tr")
                for row in table_rows:
                    try:
                        doc_type = cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(3)").text)
                        patient_name = cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(6)").text)
                        dob = cu.get_date_string(cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(7)").text)) #Assigned to Excel
                        status = cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(8)").text)
                        sent_date = cu.get_date_string(cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(9)").text))
                        document_id = cu.clean_null_data(row.find_element(By.CSS_SELECTOR, "td:nth-child(10) span.text-muted").text)

                        if status.lower()=='signed':
                            worksheet.cell(excel_row, 3).value = patient_name
                            worksheet.cell(excel_row, 15).value = doc_type
                            worksheet.cell(excel_row, 14).value=str(document_id)
                            worksheet.cell(excel_row, 8).value=cu.get_date_string(sent_date)
                            worksheet.cell(excel_row, 11).value="TRUE"
                            worksheet.cell(excel_row, 12).value="TRUE"
                            worksheet.cell(excel_row, 22).value=cu.get_date_string(dob)
                            worksheet.cell(excel_row, 17).value=rpaName
                            worksheet.cell(excel_row, 18).value=credName
                            worksheet.cell(excel_row, 19).value=location
                            excel_row=excel_row+1
                        destination_package.save(order_template)

                    except Exception as e:
                        print(str(e))
                        continue
                
                destination_package.save(order_template)
                if len(table_rows)<10:
                    break
                next_button = driver.find_element(By.XPATH, "//li[@class='page-next']/a")
                try:
                    next_button.click()
                    time.sleep(5)
                    current_page = driver.find_element(By.XPATH, "//li[@class='page-number active']/a").text
                    if first_page == current_page:
                        break
                except Exception as e:
                    break
                
                
            destination_package.save(order_template)
            time.sleep(5)
            driver.quit() 
        except Exception as e:
            print(e)
    except Exception as e:
        print(e)
        
# sendToPhysician("https://live.doctoralliance.com/Home/Login","rpaa723","rpa123","Axxess-StandardHomeHealth")